import pandas as pd
import numpy as np

INPUT_FILE = "data/state_disease_weather_master_v5.xlsx"

print("Loading dataset...")

df = pd.read_excel(INPUT_FILE)

df.columns = (
    df.columns
    .str.strip()
    .str.lower()
)

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

full_results = []

temperature_results = []
humidity_results = []
rainfall_results = []

states = sorted(
    df["state"]
    .dropna()
    .unique()
)

for state in states:

    print(f"Processing {state}")

    state_df = df[
        df["state"] == state
    ]

    diseases = sorted(
        state_df["disease"]
        .dropna()
        .unique()
    )

    for disease in diseases:

        disease_df = (
            state_df[
                state_df["disease"] == disease
            ]
            .copy()
            .sort_values(
                ["year", "month"]
            )
        )

        if len(disease_df) < 12:
            continue

        factor_best = {}

        for factor in [
            "temperature",
            "humidity",
            "rainfall"
        ]:

            lag_values = {}

            for lag in [1, 2, 3]:

                temp_df = disease_df.copy()

                temp_df["shifted_cases"] = (
                    temp_df["cases"]
                    .shift(-lag)
                )

                temp_df = temp_df.dropna(
                    subset=[
                        factor,
                        "shifted_cases"
                    ]
                )

                if len(temp_df) < 6:

                    corr = np.nan

                else:

                    corr = (
                        temp_df[factor]
                        .corr(
                            temp_df[
                                "shifted_cases"
                            ]
                        )
                    )

                lag_values[lag] = corr

                full_results.append([
                    state,
                    disease,
                    factor.title(),
                    lag,
                    corr
                ])

            valid_corrs = {
                k: v
                for k, v in lag_values.items()
                if pd.notna(v)
            }

            if not valid_corrs:
                continue

            best_lag = max(
                valid_corrs,
                key=lambda x:
                abs(valid_corrs[x])
            )

            best_corr = (
                valid_corrs[best_lag]
            )

            factor_best[factor] = [
                lag_values.get(1),
                lag_values.get(2),
                lag_values.get(3),
                best_lag,
                best_corr
            ]

        if "temperature" in factor_best:

            temperature_results.append([
                state,
                disease,
                *factor_best["temperature"]
            ])

        if "humidity" in factor_best:

            humidity_results.append([
                state,
                disease,
                *factor_best["humidity"]
            ])

        if "rainfall" in factor_best:

            rainfall_results.append([
                state,
                disease,
                *factor_best["rainfall"]
            ])

# ===================================
# DATAFRAMES
# ===================================

columns = [
    "State",
    "Disease",
    "Lag1",
    "Lag2",
    "Lag3",
    "Best_Lag",
    "Best_Correlation"
]

temperature_df = pd.DataFrame(
    temperature_results,
    columns=columns
)

humidity_df = pd.DataFrame(
    humidity_results,
    columns=columns
)

rainfall_df = pd.DataFrame(
    rainfall_results,
    columns=columns
)

full_df = pd.DataFrame(
    full_results,
    columns=[
        "State",
        "Disease",
        "Weather_Factor",
        "Lag",
        "Correlation"
    ]
)

# ===================================
# SAVE FILES
# ===================================

temperature_df.to_excel(
    "data/state_disease_temperature_lag_analysis_v5.xlsx",
    index=False
)

humidity_df.to_excel(
    "data/state_disease_humidity_lag_analysis_v5.xlsx",
    index=False
)

rainfall_df.to_excel(
    "data/state_disease_rainfall_lag_analysis_v5.xlsx",
    index=False
)

full_df.to_excel(
    "data/state_disease_lag_analysis_full_v5.xlsx",
    index=False
)

print("\nFiles Saved:")

print(
    "state_disease_temperature_lag_analysis_v5.xlsx"
)

print(
    "state_disease_humidity_lag_analysis_v5.xlsx"
)

print(
    "state_disease_rainfall_lag_analysis_v5.xlsx"
)

print(
    "state_disease_lag_analysis_full_v5.xlsx"
)

# files = [
#     "data/state_disease_temperature_lag_analysis_v5.xlsx",
#     "data/state_disease_humidity_lag_analysis_v5.xlsx",
#     "data/state_disease_rainfall_lag_analysis_v5.xlsx",
#     "data/state_disease_lag_analysis_full_v5.xlsx"
# ]

# green_fill = PatternFill(
#     fill_type="solid",
#     start_color="C6EFCE",
#     end_color="C6EFCE"
# )

# for file in files:

#     wb = load_workbook(file)

#     ws = wb.active

#     # for row in range(2, ws.max_row + 1):

#     #     for cell in ws[row]:
#     for row in range(2, ws.max_row + 1):

#             # Only Correlation Columns

#             for col in [3, 4, 5, 7]:

#                 cell = ws.cell(
#                     row=row,
#                     column=col
#                 )

#                 value = cell.value

#                 if (
#                     isinstance(
#                         value,
#                         (int, float)
#                     )
#                     and abs(value) >= 0.55
#                 ):
#                     cell.fill = green_fill

#             value = cell.value

#             if (
#                 isinstance(
#                     value,
#                     (int, float)
#                 )
#                 and abs(value) >= 0.55
#             ):
#                 cell.fill = green_fill

#     wb.save(file)

#     print(
#         f"Formatted: {file}"
#     )

files = [
    "data/state_disease_temperature_lag_analysis_v5.xlsx",
    "data/state_disease_humidity_lag_analysis_v5.xlsx",
    "data/state_disease_rainfall_lag_analysis_v5.xlsx"
]

light_green = PatternFill(
    fill_type="solid",
    start_color="C6EFCE",
    end_color="C6EFCE"
)

dark_green = PatternFill(
    fill_type="solid",
    start_color="70AD47",
    end_color="70AD47"
)

for file in files:

    wb = load_workbook(file)

    ws = wb.active

    for row in range(2, ws.max_row + 1):

        # Only Lag1, Lag2, Lag3, Best_Correlation

        # for col in [3, 4, 5, 7]:

        #     cell = ws.cell(
        #         row=row,
        #         column=col
        #     )

        #     value = cell.value

        #     if not isinstance(
        #         value,
        #         (int, float)
        #     ):
        #         continue

        #     if abs(value) >= 0.75:

        #         cell.fill = dark_green

        #     elif abs(value) >= 0.55:

        #         cell.fill = light_green

        for row in range(2, ws.max_row + 1):

            highlight = None

            for col in [3, 4, 5, 7]:

                value = ws.cell(
                    row=row,
                    column=col
                ).value

                if not isinstance(
                    value,
                    (int, float)
                ):
                    continue

                if abs(value) >= 0.75:

                    highlight = dark_green
                    break

                elif abs(value) >= 0.55:

                    highlight = light_green

            if highlight:

                for col in range(
                    1,
                    ws.max_column + 1
                ):

                    ws.cell(
                        row=row,
                        column=col
                    ).fill = highlight

    wb.save(file)

    print(
        f"Formatted: {file}"
    )

# =====================================
# FULL ANALYSIS FILE
# =====================================

wb = load_workbook(
    "data/state_disease_lag_analysis_full_v5.xlsx"
)

ws = wb.active

for row in range(2, ws.max_row + 1):

    value = ws.cell(
        row=row,
        column=5
    ).value

    if not isinstance(
        value,
        (int, float)
    ):
        continue

    if abs(value) >= 0.75:

        fill = dark_green

    elif abs(value) >= 0.55:

        fill = light_green

    else:

        continue

    # Highlight entire row

    for col in range(
        1,
        ws.max_column + 1
    ):

        ws.cell(
            row=row,
            column=col
        ).fill = fill

wb.save(
    "data/state_disease_lag_analysis_full_v5.xlsx"
)

print(
    "Formatted: state_disease_lag_analysis_full_v5.xlsx"
)