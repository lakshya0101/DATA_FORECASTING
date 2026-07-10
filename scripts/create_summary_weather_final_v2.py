import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

INPUT_FILE = "data/State_Disease_Reference_Analysis_v4.xlsx"
OUTPUT_FILE = "data/SUMMARY_WEATHER_FINAL_V2.xlsx"

print("Loading workbook...")

xls = pd.ExcelFile(INPUT_FILE)

master_df = pd.read_excel(
    "data/state_disease_weather_master.xlsx"
)

master_df.columns = (
    master_df.columns
    .str.strip()
    .str.lower()
)

rows = []

for state in xls.sheet_names:

    print(f"Processing {state}")

    df = pd.read_excel(
        INPUT_FILE,
        sheet_name=state,
        header=None
    )

    disease_rows = df[
        df[0].astype(str).str.startswith("DISEASE :", na=False)
    ].index.tolist()

    for start_row in disease_rows:

        disease_name = (
            str(df.iloc[start_row, 0])
            .replace("DISEASE :", "")
            .strip()
        )

        weather_df = master_df[
            (master_df["state"] == state)
            &
            (master_df["disease"] == disease_name)
        ]

        if len(weather_df) < 6:
            continue

        # avg_temp = round(
        #     weather_df["temperature"].mean(),
        #     3
        # )

        # avg_hum = round(
        #     weather_df["humidity"].mean(),
        #     3
        # )

        # avg_rain = round(
        #     weather_df["rainfall"].mean(),
        #     3
        # )

        avg_temp = round(
            weather_df["temperature"].mean(),
            3
        )

        avg_hum = round(
            weather_df["humidity"].mean(),
            3
        )

        avg_rain = round(
            weather_df["rainfall"].mean(),
            3
        )

        print(
            state,
            "|",
            disease_name,
            "| TEMP =",
            avg_temp,
            "| HUM =",
            avg_hum,
            "| RAIN =",
            avg_rain
        )

        lag_start = None

        for r in range(
            start_row,
            min(start_row + 120, len(df))
        ):

            value = str(df.iloc[r, 0]).strip()

            if value == "Lag Analysis":
                lag_start = r
                break

            if "Insufficient data" in value:
                lag_start = None
                break

        if lag_start is None:
            continue

        factor_data = {
            "Rainfall": [],
            "Temperature": [],
            "Humidity": []
        }

        for r in range(
            lag_start + 2,
            min(lag_start + 20, len(df))
        ):

            factor = str(df.iloc[r, 0]).strip()

            if factor.startswith("DISEASE :"):
                break

            if factor not in factor_data:
                continue

            lag = pd.to_numeric(df.iloc[r, 1], errors="coerce")
            corr = pd.to_numeric(df.iloc[r, 2], errors="coerce")

            if pd.isna(lag) or pd.isna(corr):
                continue

            factor_data[factor].append(
                (int(lag), abs(float(corr)))
            )

        # ==================================
        # WEATHER VALUES FROM MASTER DATA
        # ==================================

        master_df = pd.read_excel(
            "data/state_disease_weather_master.xlsx"
        )

        master_df.columns = (
            master_df.columns
            .str.strip()
            .str.lower()
        )

        weather_df = master_df[
            (master_df["state"] == state)
            &
            (master_df["disease"] == disease_name)
        ]

        avg_temp = round(
            weather_df["temperature"].mean(),
            3
        ) if not weather_df.empty else None

        avg_hum = round(
            weather_df["humidity"].mean(),
            3
        ) if not weather_df.empty else None

        avg_rain = round(
            weather_df["rainfall"].mean(),
            3
        ) if not weather_df.empty else None


        # for factor in ["Rainfall", "Temperature", "Humidity"]:

        #     if not factor_data[factor]:
        #         continue

        #     best_lag, best_corr = max(
        #         factor_data[factor],
        #         key=lambda x: x[1]
        #     )

        #     rows.append([
        #         state,
        #         disease_name,

        #         avg_temp if factor == "Temperature" else "",
        #         avg_rain if factor == "Rainfall" else "",
        #         avg_hum if factor == "Humidity" else "",

        #         round(best_corr, 4),

        #         best_lag
        #     ])

        for factor in ["Rainfall", "Temperature", "Humidity"]:

            if not factor_data[factor]:
                continue

            best_lag, best_corr = max(
                factor_data[factor],
                key=lambda x: x[1]
            )

            rows.append([
                state,
                disease_name,

                avg_temp
                if factor == "Temperature"
                else "",

                avg_rain
                if factor == "Rainfall"
                else "",

                avg_hum
                if factor == "Humidity"
                else "",

                round(best_corr, 4),

                best_lag
            ])

        

        # for factor in ["Rainfall", "Temperature", "Humidity"]:

        #     if not factor_data[factor]:
        #         continue

        #     best_lag, best_corr = max(
        #         factor_data[factor],
        #         key=lambda x: x[1]
        #     )

        #     rows.append([
        #         state,
        #         disease_name,
        #         best_corr if factor == "Temperature" else "",
        #         best_corr if factor == "Rainfall" else "",
        #         best_corr if factor == "Humidity" else "",
        #         best_corr,
        #         # best_lag
        #         f"{factor} ({best_lag})"
        #     ])

summary_df = pd.DataFrame(
    rows,
    columns=[
        "STATE",
        "CORELL_1_DISEASE",
        "TEMPERATURE",
        "RAINFALL",
        "HUMIDITY",
        "CORRELATION",
        "LAG_ANALYSIS"
    ]
)

with pd.ExcelWriter(
    OUTPUT_FILE,
    engine="openpyxl"
) as writer:

    summary_df.to_excel(
        writer,
        sheet_name="SUMMARY",
        startrow=2,
        index=False
    )

wb = load_workbook(OUTPUT_FILE)
ws = wb["SUMMARY"]

ws.merge_cells("A1:G1")
ws["A1"] = "SUMMARY"

ws["A1"].font = Font(
    bold=True,
    size=14
)

ws["A1"].alignment = Alignment(
    horizontal="center"
)

thin = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

for row in ws.iter_rows():
    for cell in row:
        cell.border = thin
        cell.alignment = Alignment(
            horizontal="center"
        )

for col in ws.columns:
    width = 15
    for cell in col:
        try:
            width = max(
                width,
                len(str(cell.value)) + 2
            )
        except:
            pass

    ws.column_dimensions[
        get_column_letter(col[0].column)
    ].width = width

wb.save(OUTPUT_FILE)

print(f"Saved: {OUTPUT_FILE}")
